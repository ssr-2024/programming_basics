# Step 1: Import the necessary libraries

import openpyxl as px
from pathlib import Path

# Step 2: Define the paths to the input and output files
input_path = Path('code/final_report/Beispieldaten_Tutorial.xlsx')
output_path = Path('code/final_report/Beispieldaten_Ergebnis.xlsx')

# Step 3: Load the workbook and select the active worksheet
workbook = px.load_workbook(input_path)
sheet = workbook.active

# Step 4: Rename the active worksheet to "Gesamtdaten"
sheet.title = "Gesamtdaten"

# Step 5: Define the column to evaluate (club names)
verein_spalte = 4  # Column D is the 4th column
max_col = sheet.max_column
max_row = sheet.max_row

# List of known clubs
known_clubs = [
    "FC Alle", "FC Attiswil", "FC Azzurri Bienne", "FC Bremgarten", "FC Bülach", 
    "FC Celerina", "FC Dielsdorf", "FC Erguël", "FC Frenkendorf", "FC Frutigen", 
    "FC La Combe", "FC Lutry", "FC Münchwilen", "FC Münsingen", "FC Muri", 
    "FC Sarine-Ouest", "FC Savièse", "FC Volketswil", "FC Wallisellen", 
    "FC Wetzikon", "FFV Basel", "SC Cham", "SC Düdingen"
]

# Collect data per club
vereinsdaten = {club: [] for club in known_clubs}
for row in range(2, max_row + 1):
    verein = sheet.cell(row=row, column=verein_spalte).value
    if verein in vereinsdaten:
        # Collect all data in the row for the club
        vereinsdaten[verein].append([sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)])

# Step 6: Create a new worksheet for each club
for verein, daten in vereinsdaten.items():
    # Create a new sheet with the club name 
    new_sheet = workbook.create_sheet(title=verein)
    # Copy the header
    for col in range(1, max_col + 1):
        new_sheet.cell(row=1, column=col).value = sheet.cell(row=1, column=col).value
    # Insert the data
    for row_idx, row_data in enumerate(daten, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            new_sheet.cell(row=row_idx, column=col_idx).value = value

# Step 7: Save the workbook
workbook.save(output_path)

print(f"Ergebnisdatei gespeichert unter: {output_path}")


