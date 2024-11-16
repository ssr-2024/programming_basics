# Step 1: Import the necessary libraries
import openpyxl as px
from openpyxl.chart import BarChart, Reference, Series
from pathlib import Path

# Step 2: Define the path to the input file
input_path = Path('code/final_report/Beispieldaten_Ergebnis.xlsx')
output_path = Path('code/final_report/Beispieldaten_Ergebnis_mit_Durchschnitt und Balkendiagrammen.xlsx')

# Step 3: Load the workbook
workbook = px.load_workbook(input_path)

# Step 4: Iterate through all sheets in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # Find columns that start with "Zufriedenheit_" or "Wichtigkeit_"
    zufriedenheit_columns = []
    wichtigkeit_columns = []
    header = [cell.value for cell in sheet[1]]

    for col_idx, col_name in enumerate(header, start=1):
        if col_name and col_name.startswith("Zufriedenheit_"):
            zufriedenheit_columns.append((col_name, px.utils.get_column_letter(col_idx)))
        elif col_name and col_name.startswith("Wichtigkeit_"):
            wichtigkeit_columns.append((col_name, px.utils.get_column_letter(col_idx)))

    # Get the maximum row number
    max_row = sheet.max_row

    # Iterate through the specified columns to calculate averages
    for col_name, col_letter in zufriedenheit_columns + wichtigkeit_columns:
        col_index = px.utils.column_index_from_string(col_letter)
        values = []
        for row in range(2, max_row + 1):
            cell_value = sheet.cell(row=row, column=col_index).value
            if cell_value is not None:
                try:
                    values.append(float(cell_value))
                except ValueError:
                    continue

        # Step 5: Calculate the average if there are values
        if values:
            average = sum(values) / len(values)
            # Write the average to the end of the column
            sheet.cell(row=max_row + 1, column=col_index).value = average

    # Step 6: Function to create a bar chart
    def create_bar_chart(title, columns, position):
        if not columns:
            return  # Skip if there are no matching columns

        chart = BarChart()
        chart.title = title
        chart.y_axis.title = "Durchschnittswerte"
        chart.y_axis.scaling.min = 1
        chart.y_axis.scaling.max = 5
        chart.legend.position = 'b'  # Position the legend below the chart

        for col_name, col_letter in columns:
            # Use the last value in each column for the chart
            data = Reference(sheet, min_col=px.utils.column_index_from_string(col_letter), min_row=max_row + 1, max_row=max_row + 1)
            series = Series(data, title=col_name)  # Use column names as legend titles
            chart.series.append(series)
            chart.set_categories(Reference(sheet, min_col=px.utils.column_index_from_string(col_letter), min_row=1, max_row=1))

        # Place the chart on the sheet
        sheet.add_chart(chart, position)

    # Step 7: Create charts for "Zufriedenheit" and "Wichtigkeit"
    create_bar_chart("Zufriedenheit", zufriedenheit_columns, f"A{max_row + 3}")
    create_bar_chart("Wichtigkeit", wichtigkeit_columns, f"K{max_row + 3}")

# Step 8: Save the workbook with the charts and averages
workbook.save(output_path)
print(f"Zweite Ergebnisdatei gespeichert unter: {output_path}. Herzliche Gratulation. Du hast nun beide Übungsaufgaben gelöst.")

